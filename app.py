from flask import Flask, render_template, jsonify, request
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
import numpy as np
from sklearn.linear_model import LinearRegression
from io import StringIO
import csv
from datetime import datetime

app = Flask(__name__)
CORS(app)
app.config['SECRET_KEY'] = 'student-performance-2025'
app.config['SQLALCHEMY_DATABASE_URI'] = 'postgresql://postgres:senthil777@localhost/studentdb'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

class Student(db.Model):
    __tablename__ = 'students'
    id = db.Column(db.Integer, primary_key=True)
    school = db.Column(db.String(10))
    gender = db.Column(db.String(1))
    age = db.Column(db.Integer)
    study = db.Column(db.String(20))
    failures = db.Column(db.Integer)
    absences = db.Column(db.Integer)
    g1 = db.Column(db.Integer)
    g2 = db.Column(db.Integer)
    g3 = db.Column(db.Integer)

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/api/students")
def api_students():
    students = Student.query.all()
    return jsonify([{
        "id": s.id,
        "school": s.school or "-",
        "gender": s.gender or "-",
        "age": s.age or 16,
        "study": s.study or "2-5h",
        "failures": s.failures or 0,
        "absences": s.absences or 0,
        "g1": s.g1 or 10,
        "g2": s.g2 or 10,
        "g3": s.g3 or 10
    } for s in students])

@app.route("/api/predict", methods=["POST"])
def api_predict():
    data = request.get_json()
    try:
        all_students = Student.query.all()
        if len(all_students) > 5:
            # Convert study time to numeric for ML
            def study_to_num(study):
                if '1-2' in str(study): return 1
                if '2-5' in str(study): return 2
                if '5-10' in str(study): return 3
                return 2
            
            X = np.array([[s.age or 16, study_to_num(s.study), s.g1 or 10, s.g2 or 10, 
                          s.failures or 0, s.absences or 0] for s in all_students])
            y = np.array([s.g3 or 10 for s in all_students])
            model = LinearRegression().fit(X, y)
            features = np.array([[data['age'], data['study'], data['g1'], data['g2'], 
                                data['failures'], data['absences']]])
            predicted = max(0, min(20, model.predict(features)[0]))
        else:
            # Fallback formula
            predicted = (data['g1'] * 0.3 + data['g2'] * 0.3 + data['study'] * 1 + 
                        10 - data['failures'] * 1.5 - data['absences'] * 0.2)
            predicted = max(0, min(20, round(predicted, 1)))
    except Exception as e:
        print(f"Prediction error: {e}")
        predicted = 12.5
    
    status = "Excellent" if predicted >= 15 else "Good" if predicted >= 10 else "Average"
    return jsonify({"predicted_grade": round(predicted, 1), "status": status})

@app.route("/api/export")
def api_export():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    
    students = Student.query.all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Data"
    
    # Header row with styling
    headers = ['ID', 'School', 'Gender', 'Age', 'Study Time', 'Failures', 'Absences', 'G1', 'G2', 'G3']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    for row_num, student in enumerate(students, 2):
        ws.cell(row=row_num, column=1).value = student.id
        ws.cell(row=row_num, column=2).value = student.school
        ws.cell(row=row_num, column=3).value = student.gender
        ws.cell(row=row_num, column=4).value = student.age
        ws.cell(row=row_num, column=5).value = student.study
        ws.cell(row=row_num, column=6).value = student.failures or 0
        ws.cell(row=row_num, column=7).value = student.absences or 0
        ws.cell(row=row_num, column=8).value = student.g1 or 0
        ws.cell(row=row_num, column=9).value = student.g2 or 0
        ws.cell(row=row_num, column=10).value = student.g3 or 0
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue(), 200, {
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'Content-Disposition': f'attachment; filename=student_performance_{datetime.now().strftime("%Y%m%d")}.xlsx'
    }


if __name__ == "__main__":
    with app.app_context():
        count = Student.query.count()
        print("✅ Connected to PostgreSQL studentdb!")
        print(f"✅ Found {count} students in 'students' table")
    app.run(debug=True, host='0.0.0.0', port=5000)
